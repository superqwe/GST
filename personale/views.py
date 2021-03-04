import datetime
import functools
import itertools
import os
from datetime import timedelta

import openpyxl
import pandas as pd
from django.db.models import Count
from django.http import HttpResponse, HttpResponseRedirect
from django.template import loader
from django_pandas.io import read_frame
from openpyxl.styles import Side, Border, PatternFill, Font, Alignment

import personale.viste.views_aggiorna_unilav as views_aggiorna_unilav
from personale import views_simulazione_emergenze, views_programma_visite_mediche, views_util, views_programma_officina
from personale.admin_actions import data_ultima_modifica_leggi
from personale.models import Lavoratore, Azienda
from personale.views_estrai_dati import estrazione_selettiva2, estrazione_da_excel2
from personale.views_programma_officina import N_CARD_PER_RIGO, TRONCA_NOME
from personale.views_tesserini import genera_tesserini
from personale.views_util import autorizzato

from pprint import pprint as pp

def index(request):
    return HttpResponse("Hello, world. You're at the ''personale'' index.")


def completo(request, filtro=False, ordinamento=None):
    # pagina_attiva = 'in_forza' if filtro == 'in_forza' else 'tutti'
    tabella_completa = False

    if ordinamento == 'a':
        dati = views_util.lavoratori_suddivisi_per_azienda()
        pagina_attiva = 'azienda'
    elif ordinamento == 'am':
        dati = views_util.lavoratori_suddivisi_per_azienda('arcelormittal')
        pagina_attiva = 'arcelormittal'
    elif ordinamento == 'andritz':
        dati = views_util.lavoratori_suddivisi_per_azienda('andritz')
        pagina_attiva = 'andritz'
    elif ordinamento == 'eni':
        dati = views_util.lavoratori_suddivisi_per_azienda('eni')
        pagina_attiva = 'eni'
    elif ordinamento == 'edison':
        dati = views_util.lavoratori_suddivisi_per_azienda('edison')
        pagina_attiva = 'edison'
    elif ordinamento == 'c':
        dati = views_util.lavoratori_suddivisi_per_azienda('cantiere')
        pagina_attiva = 'cantiere'
    elif ordinamento == 'n':
        dati = views_util.lavoratori_con_nomine()
        pagina_attiva = 'nomine'
    elif ordinamento == 's':
        dati = views_util.lavoratori_suddivisi_per_azienda('stato')
        pagina_attiva = 'scadenza'
    elif ordinamento == 'v':
        dati = views_util.lavoratori_suddivisi_per_azienda('idoneita')
        pagina_attiva = 'idoneita'
    else:
        dati = views_util.lavoratori_suddivisi_per_azienda(in_forza=False)
        pagina_attiva = 'tutti'

    template = loader.get_template('personale/principale.html')
    context = {
        'autorizzato': autorizzato(request.user),
        'dati': dati,
        'pagina_attiva': pagina_attiva,
        'scadenza': views_util.DateScadenza(),
        'tabella_completa': tabella_completa,
        'data_ultima_modifica': data_ultima_modifica_leggi(),
    }
    return HttpResponse(template.render(context, request))


def unilav(request):
    oggi = datetime.date.today()
    fino_al = oggi + timedelta(7)

    lavoratori = Lavoratore.objects.filter(in_forza=True, azienda=Azienda.objects.get(nome='Modomec'),
                                           unilav__lte=fino_al, unilav__gt=oggi)
    lavoratori_r = Lavoratore.objects.filter(in_forza=True, azienda=Azienda.objects.get(nome='Modomec'),
                                             unilav__lte=oggi)
    rossi = Lavoratore.objects.filter(in_forza=True, azienda=Azienda.objects.get(nome='Modomec'),
                                      unilav__lt=oggi)
    gialli = Lavoratore.objects.filter(in_forza=True, azienda=Azienda.objects.get(nome='Modomec'),
                                       unilav=oggi)
    verdi = Lavoratore.objects.filter(in_forza=True, azienda=Azienda.objects.get(nome='Modomec'),
                                      unilav__lte=fino_al, unilav__gt=oggi)

    n = {'r': rossi.count(), 'g': gialli.count(), 'v': verdi.count()}
    n['t'] = n['v'] + n['g'] + n['r']

    template = loader.get_template('personale/unilav.html')
    context = {
        'autorizzato': autorizzato(request.user),
        'fino_al': fino_al,
        'lavoratori': lavoratori,
        'lavoratori_r': lavoratori_r,
        'scadenza': views_util.DateScadenza(),
        'n': n
    }

    return HttpResponse(template.render(context, request))


def unilav_scaduti(request):
    oggi = datetime.date.today()

    lavoratori_r = Lavoratore.objects.filter(in_forza=True, azienda=Azienda.objects.get(nome='Modomec'),
                                             unilav__lt=oggi)

    print('\nLavoratori passati in non in forza:')
    for lavoratore in lavoratori_r:
        print('\t', lavoratore.cognome, lavoratore.nome)
        lavoratore.in_forza = False
        lavoratore.save()
    print()

    template = loader.get_template('personale/unilav.html')
    context = {
        'autorizzato': autorizzato(request.user),
        'lavoratori_r': lavoratori_r,
        'scadenza': views_util.DateScadenza(),
        'scaduti': True,
    }

    return HttpResponse(template.render(context, request))


def mansioni(request):
    elenco_mansioni = Lavoratore.objects.filter(azienda=Azienda.objects.get(nome='Modomec')).values(
        'mansione').annotate(totale=Count('mansione')).order_by('-totale')

    dati = []
    for mansione in elenco_mansioni:
        dati.append((mansione['totale'], mansione['mansione']))
        print('%2i' % mansione['totale'], mansione['mansione'])

    dati = pd.DataFrame(dati, columns=('n', 'mansione'))
    dati.to_excel('mansioni.xlsx', sheet_name='mansioni')

    ora = datetime.datetime.now()
    return HttpResponse("""<h1 style="text-align:center">mansioni.xlsx creato</h1>
                        <h2 style="text-align:center"> %s </h2>""" % ora)


def test(request):
    elenco_mansioni = Lavoratore.objects.filter(azienda=Azienda.objects.get(nome='Modomec')).values(
        'mansione').annotate(totale=Count('mansione')).order_by('-totale')

    for mansione in elenco_mansioni:
        print('%2i' % mansione['totale'], mansione['mansione'])

    ora = datetime.datetime.now()
    return HttpResponse("""<h1 style="text-align:center">test</h1>
                        <h2 style="text-align:center"> %s </h2>""" % ora)


def estrai_dati(request):
    opzioni_estrazione = views_util.EstraiDatiUtil()

    dati = opzioni_estrazione.estrai_documenti()

    template = loader.get_template('personale/principale.html')
    context = {
        'autorizzato': autorizzato(request.user),
        'estrai_dati': True,
        'data_ultima_modifica': data_ultima_modifica_leggi(),
        'struttura': dati['struttura'],
        'estrazione': dati['estrazione'],
        'filtro_impresa': dati['filtro_impresa'],
        'cantieri': dati['filtro_cantiere'],
        'scadenza': views_util.DateScadenza(),
    }

    return HttpResponse(template.render(context, request))


def formazione(request):
    includi_idoneita = True
    includi_unilav = True
    ora = datetime.datetime.now()
    modomec = Lavoratore.objects.filter(azienda__nome='Modomec', in_forza=True).order_by('cognome', 'nome')
    building = Lavoratore.objects.filter(azienda__nome='Building', in_forza=True).order_by('cognome', 'nome')
    rimec = Lavoratore.objects.filter(azienda__nome='Rimec', in_forza=True).order_by('cognome', 'nome')
    welding = Lavoratore.objects.filter(azienda__nome='Welding', in_forza=True).order_by('cognome', 'nome')

    lavoratori = (modomec, building, rimec, welding)
    aziende = ('modomec', 'building', 'rimec', 'welding')
    # aziende = ('modomec', 'building', 'welding')

    colonne_escluse = ['id', 'in_forza', 'azienda', 'ci', 'codice_fiscale', 'data_nascita', 'luogo_nascita',
                       'luogo_residenza', 'idoneita', 'data_assunzione', 'indeterminato', 'unilav', 'escavatore', 'rls',
                       'stato', 'rspp', 'nomina_preposto', 'nomina_antincendio', 'nomina_primo_soccorso', 'nomina_rls',
                       'nomina_aspp', 'foto', 'consegna_dpi']

    if includi_idoneita:
        colonne_escluse.remove('idoneita')
    if includi_unilav:
        colonne_escluse.remove('unilav')

    with pd.ExcelWriter('formazione.xlsx', engine='openpyxl') as writer:
        for lav, azienda in zip(lavoratori, aziende):
            dati = read_frame(lav)
            dati.drop(colonne_escluse, axis=1, inplace=True)
            dati.to_excel(writer, azienda)

        writer.save()

    wb = openpyxl.load_workbook('formazione.xlsx')

    for azienda in aziende:
        ws = wb[azienda]

        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A3

        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToHeight = False

        ws.print_title_rows = '1:2'
        ws.oddFooter.right.text = "Page &[Page] of &N"
        ws.oddFooter.left.text = "Aggiornato al %s " % datetime.datetime.now().strftime("%d/%m/%y")

        ws.delete_cols(1, 1)
        ws.insert_rows(1)
        ws.cell(row=1, column=1).value = azienda.upper()
        ws['A1'].font = Font(size=18, color='007e60')

        for cell in ws['A2:U2'][0]:
            cell.border = Border(top=Side(border_style='thin', color='007e60'),
                                 bottom=Side(border_style='thin', color='007e60'))

        max_row = ws.max_row
        rows = ws['A3:U%i' % max_row]
        for i, row in enumerate(rows):

            for cell in row:
                cell.border = Border(bottom=Side(border_style='thin', color='91ffe3'))
                cell.number_format = 'DD/MM/YY'

                if cell.column >= 5:  # colonna E
                    cell.alignment = Alignment(horizontal='center')

                if i % 2:
                    cell.fill = PatternFill(start_color='ebfffa', end_color='ebfffa', fill_type='solid')

        for cell in ws['A%i:V%i' % (max_row, max_row)][0]:
            cell.border = Border(bottom=Side(border_style='thin', color='000000'), )

        # todo: da sistemare larghezza automantica colonna
        # dims = {}
        # for row in ws.rows:
        #     for cell in row:
        #         if cell.value:
        #             dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
        # for col, value in dims.items():
        #     ws.column_dimensions[col].width = value
        ### fine todo

        wb.save('formazione.xlsx')

    wb.close()

    if os.path.exists('formazione.xlsx'):
        with open('formazione.xlsx', "rb") as excel:
            data = excel.read()

        response = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=%s_Report.xlsx' % id
        return response

    return HttpResponse("""<h1 style="text-align:center">formazione</h1>
                        <h2 style="text-align:center"> %s </h2>""" % ora)


def dati_estratti(request):
    # for x in request.POST:
    #     print(x, '-->', request.POST[x])

    opzioni_estrazione = views_util.EstraiDatiUtil()

    opzioni_estrazione.scrivi_cfg(request.POST)

    dati = opzioni_estrazione.estrai_documenti(True)

    imprese = [impresa[0] for impresa in dati['filtro_impresa'] if impresa[1]]
    cantieri = [cantiere[0] for cantiere in dati['filtro_cantiere'] if cantiere[1]]

    elenco_doc = []
    for classe_doc in dati['struttura']:

        for doc in classe_doc[1]:

            if doc[1]:
                elenco_doc.append(doc[0])

    dati = opzioni_estrazione.estrai_documenti(True)

    xlsx = dati['estrazione']['nome_file_xlsx']

    if dati['estrazione']['tipo_estrazione'] == 'filtri':
        lavoratori = estrazione_selettiva2(aziende=imprese, cantieri=cantieri, documenti=elenco_doc)
        tipo_estrazione = 'filtri'
    else:
        lavoratori = estrazione_da_excel2(xlsx, documenti=elenco_doc)
        tipo_estrazione = 'excel'

        if not xlsx:
            return HttpResponseRedirect('/personale/estrai_dati2/')

    dati = opzioni_estrazione.estrai_documenti(True)

    if dati['estrazione']['tipo_estrazione'] == 'filtri':
        lavoratori = estrazione_selettiva2(aziende=imprese, cantieri=cantieri, documenti=elenco_doc)
    else:
        xlsx = dati['estrazione']['nome_file_xlsx']
        lavoratori = estrazione_da_excel2(xlsx, documenti=elenco_doc)

    template = loader.get_template('personale/principale.html')
    context = {'autorizzato': autorizzato(request.user),
               'data_ultima_modifica': data_ultima_modifica_leggi(),
               'dati_estratti': True,
               'res_imprese': imprese,
               'res_cantieri': cantieri,
               'res_elenco_doc': elenco_doc,
               'res_lavoratori': lavoratori,
               'scadenza': views_util.DateScadenza(),
               'tipo_estrazione': tipo_estrazione,
               'nome_file_xlsx': xlsx,
               }

    return HttpResponse(template.render(context, request))


def aggiorna_unilav(request):
    wb = openpyxl.load_workbook('c:/Users/leonardo.masi/Desktop/UNILAV.xlsx')

    fogli = wb.sheetnames

    context = {'autorizzato': autorizzato(request.user),
               'fogli': fogli, }

    # get = foglio_selezionato = False
    if request.GET:
        get = True
        foglio_selezionato = request.GET['foglio']
        foglio = wb.active

        errore = []
        assunzione = []
        cessazione = []
        proroga = []
        trasformazione = []
        for colonna in range(1, 11, 3):
            try:
                categoria = foglio.cell(row=1, column=colonna).value.split()[0].strip().lower()
                print('\n' * 2, categoria.upper())
            except AttributeError:
                break

            rigo = 2
            while True:
                try:
                    cella = foglio.cell(row=rigo, column=colonna).value
                    cognome, nome = cella.split(maxsplit=1)
                except ValueError:
                    print('*** Errore -->', cella, '--> procedere a mano')
                    errore.append((cella, '%s: procedere a mano' % categoria.title()))
                    rigo += 3
                    continue
                except AttributeError:
                    # print('-->', cella, '<--')
                    print()
                    break

                cognome = cognome.title()
                nome = nome.title()
                cf = foglio.cell(row=rigo + 1, column=colonna).value.split(':')[1]

                print('%-15s %-15s %s' % (cognome, nome, cf))

                try:
                    data_assunzione = foglio.cell(row=rigo + 1, column=colonna + 1).value.split()[-1]
                except AttributeError:
                    print('*** Errore -->', cella, '--> comunicazione annullata')
                    errore.append((cella, '%s: comunicazione annullata' % categoria.title()))
                    rigo += 3
                    continue

                data_assunzione = datetime.datetime.strptime(data_assunzione, '%d/%m/%Y').date()

                try:
                    data_fine = foglio.cell(row=rigo + 2, column=colonna + 1).value.split()[-1]
                    data_fine = datetime.datetime.strptime(data_fine, '%d/%m/%Y').date()
                except AttributeError:
                    data_fine = None

                if categoria == 'proroga':
                    proroga.append((cognome, nome, cf, data_assunzione, data_fine))
                elif categoria == 'assunzione':
                    assunzione.append((cognome, nome, cf, data_assunzione, data_fine))
                elif categoria == 'cessazione':
                    cessazione.append((cognome, nome, cf, data_assunzione, data_fine))
                elif categoria == 'trasformazione':
                    trasformazione.append((cognome, nome, cf, data_assunzione, data_fine))

                rigo += 3

        assunzione.sort()
        cessazione.sort()
        proroga.sort()
        trasformazione.sort()

        f_cessazione = [x[:3] for x in cessazione]
        f_assunzione = [x for x in assunzione if x[:3] not in f_cessazione]
        proroga = [x for x in proroga if x[:3] not in f_cessazione]
        trasformazione = [x for x in trasformazione if x[:3] not in f_cessazione]

        errori_assunzione = views_aggiorna_unilav.assunzione(f_assunzione)
        errori_proroga = views_aggiorna_unilav.proroga(proroga)
        errori_trasformazione = views_aggiorna_unilav.trasformazione(trasformazione)
        errori_cessazione = views_aggiorna_unilav.cessazione(cessazione)

        errore.extend(errori_assunzione)
        errore.extend(errori_proroga)
        errore.extend(errori_trasformazione)
        errore.extend(errori_cessazione)
        errore.sort()

        context = {'autorizzato': autorizzato(request.user),
                   'data_ultima_modifica': data_ultima_modifica_leggi(),
                   'fogli': fogli,
                   'foglio_selezionato': foglio_selezionato,
                   'get': get,
                   'dati': ((proroga, assunzione, cessazione, trasformazione),),
                   'errori': errore,
                   'scrivi': assunzione,
                   }

    template = loader.get_template('personale/aggiorna_unilav.html')

    return HttpResponse(template.render(context, request))


def rait(request):
    context = {'autorizzato': autorizzato(request.user),
               'data_ultima_modifica': data_ultima_modifica_leggi(),
               }

    template = loader.get_template('personale/rait.html')
    return HttpResponse(template.render(context, request))


def rait_estratti(request):
    context = {'autorizzato': autorizzato(request.user),
               'data_ultima_modifica': data_ultima_modifica_leggi(),
               'estratti': 'True',
               }

    template = loader.get_template('personale/rait.html')
    return HttpResponse(template.render(context, request))


def programma_officina(request):
    elenco_lavoratori, righe, (settimana, dal, al, rev) = views_programma_officina.programma_officina()

    max_width_card = 100 // N_CARD_PER_RIGO
    context = {'autorizzato': autorizzato(request.user),
               'elenco_lavoratori': elenco_lavoratori,
               'righe': righe,
               'contatore_lavoratori': functools.partial(next, itertools.count(1)),
               'tronca_nome': TRONCA_NOME,
               'mw': max_width_card,
               'programma_dal': dal,
               'programma_al': al,
               'programma_rev': rev,
               'settimana': settimana,
               }

    template = loader.get_template('personale/programma_officina/programma_officina.html')
    return HttpResponse(template.render(context, request))


def tesserini(request):
    genera_tesserini()
    ora = datetime.datetime.now()
    return HttpResponse("""<h1 style="text-align:center">tesserini creato</h1>
                        <h2 style="text-align:center"> %s </h2>""" % ora)


def simulazione_emergenze(request, filtro=None):
    context = views_simulazione_emergenze.simulazione_emergenze(request)

    template = loader.get_template('personale/simulazioni_emergenze/simulazioni_emergenze.html')
    return HttpResponse(template.render(context, request))


def programma_visite_mediche(request):
    ora = datetime.datetime.now()
    fout = views_programma_visite_mediche.programma_visite_mediche()

    if os.path.exists(fout):
        with open(fout, "rb") as excel:
            data = excel.read()

        response = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=%s_Visite.xlsx' % id
        return response

    return HttpResponse("""<h1 style="text-align:center">Programma Visite Mediche</h1>
                        <h2 style="text-align:center"> %s </h2>""" % ora)
