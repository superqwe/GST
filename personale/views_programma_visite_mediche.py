import datetime

import openpyxl
from django.db.models import Q
from django_pandas.io import read_frame
from openpyxl.styles import Border, Side, Alignment

from personale.models import Lavoratore
import pandas as pd

FOUT = 'visite_mediche.xlsx'
GIORNI_SCADENZA = 30


def programma_visite_mediche():
    ora = datetime.datetime.now()
    lavoratori = Lavoratore.objects \
        .filter(Q(idoneita=None) | Q(idoneita__lte=ora + datetime.timedelta(days=GIORNI_SCADENZA)), in_forza=True) \
        .exclude(Q(azienda__nome='-')
                 | Q(cantiere__nome='Andritz (DE)')
                 | Q(cantiere__nome='Andritz (NL)')
                 | Q(cantiere__nome='Ansaldo (VE)')
                 | Q(cantiere__nome='ArcelorMittal')
                 | Q(cantiere__nome='Fincantieri (AN)')
                 | Q(cantiere__nome='Fincantieri (GO)')
                 | Q(cantiere__nome='ISAB (SR)')
                 ) \
        .order_by('azienda', 'cantiere', 'cognome', 'nome')

    with pd.ExcelWriter(FOUT, engine='openpyxl') as writer:
        pd_dati = read_frame(lavoratori)
        dati = pd_dati[['azienda', 'cantiere', 'cognome', 'nome', 'mansione', 'idoneita']]
        dati.to_excel(writer)
        writer.save()

    wb = openpyxl.load_workbook(FOUT)
    ws = wb['Sheet1']

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.delete_cols(1, 1)

    max_row = ws.max_row
    rows = ws['A2:F%i' % max_row]
    for row in rows:

        for cell in row:
            cell.border = Border(bottom=Side(border_style='thin', color='91ffe3'))

            if cell.column == 'F':
                cell.number_format = 'DD/MM/YY'
                cell.alignment = Alignment(horizontal='center')

    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))

    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    for cell in ws['A%i:F%i' % (max_row, max_row)][0]:
        cell.border = Border(bottom=Side(border_style='thin', color='000000'), )

    wb.save(FOUT)
    wb.close()

    return FOUT
